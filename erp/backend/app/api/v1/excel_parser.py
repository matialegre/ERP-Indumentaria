"""
Parser de Excel de Notas de Pedido (portado de CONTROL REMITOS).
Soporta:
  - Miding / Montagne / World Sport (formato genérico con CODALFA + CANTIDAD)
  - OMBAK (multi-solapa con packs y filas ocultas)
"""
import re
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, UploadFile, HTTPException
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.api.deps import get_current_user
from app.models.user import User
from app.models.provider import Provider

router = APIRouter(prefix="/excel-parser", tags=["Excel Parser"])


def _safe_int(v):
    if v is None:
        return 0
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def _safe_float(v):
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _parse_curva_distribution(curva_text: str) -> dict:
    if not curva_text:
        return {}
    s = str(curva_text).strip()
    result = {}
    parts = re.split(r'[,\-]+', s)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        m = re.match(r'^(.+?)\s*[xX]\s*(\d+)$', part)
        if m:
            label = m.group(1).strip()
            count = int(m.group(2))
            result[label] = count
    return result


def _parsear_excel_ombak(file_bytes: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    solapas = []
    items_por_codigo = {}
    total_unidades_global = 0
    total_items_global = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "resumen":
            continue
        ws = wb[sheet_name]

        hidden_rows = set()
        for row_idx in range(1, ws.max_row + 1):
            rd = ws.row_dimensions.get(row_idx)
            if rd and rd.hidden:
                hidden_rows.add(row_idx)

        header_row = None
        col_map = {}
        for r in range(1, min(12, ws.max_row + 1)):
            for c in range(1, min(20, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if not v:
                    continue
                vu = str(v).upper().strip().replace('_', ' ')
                if vu == 'CODIGO' and not header_row:
                    header_row = r
                    col_map['codigo'] = c
            if header_row:
                break

        if not header_row:
            solapas.append({"nombre": sheet_name, "items": [], "total_unidades": 0, "total_items": 0, "sin_datos": True})
            continue

        for c in range(1, min(20, ws.max_column + 1)):
            v = ws.cell(header_row, c).value
            if not v:
                continue
            vu = str(v).upper().strip().replace('_', ' ')
            if 'MODELO' in vu or 'ARTICULO' in vu:
                col_map['modelo'] = c
            elif vu in ('CANTIDAD', 'CANTIDAD PACKS'):
                col_map['cantidad'] = c
            elif vu in ('TOTAL UNDS', 'TOTAL UNIDS', 'TOTAL PARES'):
                col_map['total_unds'] = c
            elif 'PACK PEDIDO' in vu:
                col_map['pack_pedido'] = c
            elif vu == 'PACK':
                col_map['pack_size'] = c
            elif 'CURVA' in vu or 'DESCRIPCION' in vu:
                col_map['curva'] = c
            elif vu == 'TALLE':
                col_map['talle'] = c
            elif 'COLOR' in vu or 'ESPECIFICACION' in vu:
                col_map['color'] = c
            elif 'PRECIO MAYORISTA' in vu or vu == 'PRECIO UNITARIO':
                col_map['precio'] = c
            elif vu == 'CATEGORIA':
                col_map['categoria'] = c

        def _cell(r, role):
            c = col_map.get(role)
            return ws.cell(r, c).value if c else None

        def _code_at(r):
            v = _cell(r, 'codigo')
            if not v:
                return None
            s = str(v).strip().rstrip('.')
            return s if re.match(r'^\d{2}/\d{3,5}$', s) else None

        all_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            code = _code_at(r)
            if not code:
                continue
            is_hidden = r in hidden_rows
            all_rows.append({
                'row': r,
                'codigo': code,
                'modelo': str(_cell(r, 'modelo') or '').strip(),
                'curva': str(_cell(r, 'curva') or _cell(r, 'color') or '').strip(),
                'talle': str(_cell(r, 'talle') or '').strip(),
                'pack_size': _safe_int(_cell(r, 'pack_size')),
                'pack_pedido': _safe_int(_cell(r, 'pack_pedido')),
                'total_unds': _safe_int(_cell(r, 'total_unds')),
                'cantidad': _safe_int(_cell(r, 'cantidad')),
                'precio': _safe_float(_cell(r, 'precio')),
                'categoria': str(_cell(r, 'categoria') or '').strip(),
                'hidden': is_hidden,
            })

        blocks = []
        i = 0
        while i < len(all_rows):
            row = all_rows[i]
            if not row['hidden']:
                children = []
                j = i + 1
                while j < len(all_rows) and all_rows[j]['hidden']:
                    children.append(all_rows[j])
                    j += 1
                blocks.append((row, children))
                i = j
            else:
                i += 1

        items = []
        for parent, children in blocks:
            pack_ped = parent['pack_pedido']
            pack_sz = parent['pack_size']
            total_u = parent['total_unds']
            qty_direct = parent['cantidad']
            curva = parent['curva']

            if children:
                if pack_ped <= 0 and total_u <= 0 and qty_direct <= 0:
                    continue
                child_codes = set(ch['codigo'] for ch in children)
                parent_code_in_children = parent['codigo'] in child_codes
                children_have_qty = any(
                    ch['pack_size'] > 0 or ch['cantidad'] > 0 or ch['total_unds'] > 0
                    for ch in children
                )
                curva_dist = _parse_curva_distribution(curva)
                emitted_from_children = []
                if children_have_qty and (len(child_codes) > 1 or not parent_code_in_children):
                    for ch in children:
                        ch_qty = 0
                        if ch['pack_size'] > 0 and pack_ped > 0:
                            ch_qty = ch['pack_size'] * pack_ped
                        elif ch['cantidad'] > 0:
                            ch_qty = ch['cantidad']
                        if ch_qty > 0:
                            emitted_from_children.append({
                                'codigo': ch['codigo'],
                                'modelo': ch['modelo'] or parent['modelo'],
                                'precio': ch['precio'] or parent['precio'],
                                'cantidad': ch_qty,
                                'categoria': ch['categoria'] or parent['categoria'] or sheet_name,
                            })
                elif children_have_qty and parent_code_in_children:
                    for ch in children:
                        ch_qty = ch['cantidad'] or ch['total_unds'] or ch['pack_size']
                        if ch_qty <= 0 and pack_ped > 0 and ch['pack_size'] > 0:
                            ch_qty = ch['pack_size'] * pack_ped
                        if ch_qty > 0:
                            emitted_from_children.append({
                                'codigo': ch['codigo'],
                                'modelo': ch['modelo'] or parent['modelo'],
                                'precio': ch['precio'] or parent['precio'],
                                'cantidad': ch_qty,
                                'categoria': ch['categoria'] or parent['categoria'] or sheet_name,
                            })
                elif not children_have_qty and curva_dist and len(children) > 0:
                    assigned_labels = set()
                    for ch in children:
                        ch_name = ch['modelo'].upper().strip()
                        for label, cnt in curva_dist.items():
                            lu = label.upper()
                            if lu == ch_name or lu in ch_name or ch_name in lu:
                                ch_qty = cnt * max(pack_ped, 1)
                                emitted_from_children.append({
                                    'codigo': ch['codigo'],
                                    'modelo': ch['modelo'] or parent['modelo'],
                                    'precio': ch['precio'] or parent['precio'],
                                    'cantidad': ch_qty,
                                    'categoria': ch['categoria'] or parent['categoria'] or sheet_name,
                                })
                                assigned_labels.add(lu)
                                break
                    parent_qty = 0
                    for label, cnt in curva_dist.items():
                        if label.upper() not in assigned_labels:
                            parent_qty += cnt * max(pack_ped, 1)
                    if parent_qty > 0:
                        emitted_from_children.append({
                            'codigo': parent['codigo'],
                            'modelo': parent['modelo'],
                            'precio': parent['precio'],
                            'cantidad': parent_qty,
                            'categoria': parent['categoria'] or sheet_name,
                        })
                if not emitted_from_children and total_u > 0:
                    emitted_from_children.append({
                        'codigo': parent['codigo'],
                        'modelo': parent['modelo'],
                        'precio': parent['precio'],
                        'cantidad': total_u,
                        'categoria': parent['categoria'] or sheet_name,
                    })
                items.extend(emitted_from_children)
            else:
                qty = qty_direct or total_u
                if qty <= 0 and pack_ped > 0 and pack_sz > 0:
                    qty = pack_ped * pack_sz
                if qty > 0:
                    items.append({
                        'codigo': parent['codigo'],
                        'modelo': parent['modelo'],
                        'precio': parent['precio'],
                        'cantidad': qty,
                        'categoria': parent['categoria'] or sheet_name,
                    })

        for item in items:
            code = item['codigo']
            if code in items_por_codigo:
                items_por_codigo[code]['cantidad'] += item['cantidad']
            else:
                items_por_codigo[code] = dict(item)

        total_u = sum(i['cantidad'] for i in items)
        total_unidades_global += total_u
        total_items_global += len(items)
        solapas.append({
            'nombre': sheet_name,
            'items': items,
            'total_unidades': total_u,
            'total_items': len(items),
            'sin_datos': len(items) == 0,
        })

    wb.close()
    return {
        'solapas': solapas,
        'total_unidades': total_unidades_global,
        'total_items': total_items_global,
    }


def _parsear_excel_generico(file_bytes: bytes, proveedor: str = "", es_reposicion: bool = False) -> dict:
    import pandas as pd
    es_miding_repo = 'MIDING' in proveedor.upper() and es_reposicion
    try:
        xl = pd.ExcelFile(BytesIO(file_bytes))
        sheet_names = xl.sheet_names
        xl.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    hojas = []
    total_unidades_global = 0
    total_items_global = 0
    hojas_a_leer = sheet_names
    if any(s.upper() == 'PEDIDO' for s in sheet_names):
        hojas_a_leer = [s for s in sheet_names if s.upper() == 'PEDIDO']

    for sheet in hojas_a_leer:
        upper = sheet.upper()
        if any(k in upper for k in ('RESUMEN', 'TOTAL', 'CONSOLIDADO', 'RUTA', 'PLANTILLA')):
            continue
        try:
            raw_df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None)
            header_row = None
            for i, row in raw_df.iterrows():
                row_str = ' '.join([str(x) for x in row.values if pd.notna(x)]).upper()
                if any(k in row_str for k in ['TOTAL PARES', 'CODALFA', 'COD_ALFA', 'CANTIDAD', 'CANT. TOTAL', 'CANT TOTAL', 'ARTÍCULO']):
                    header_row = i
                    break

            headerless = False
            if header_row is None and len(raw_df.columns) >= 6:
                sample = raw_df.head(min(5, len(raw_df)))
                try:
                    col0_alpha = sample[0].apply(lambda x: isinstance(x, str) and len(x) >= 6 and x[:2].isalpha()).all()
                    col5_numeric = sample[5].apply(lambda x: isinstance(x, (int, float)) and 0 < x < 1000).all()
                    col3_alpha = sample[3].apply(lambda x: isinstance(x, str) and len(x) >= 10).all()
                    if col0_alpha and col5_numeric and col3_alpha:
                        headerless = True
                except Exception:
                    headerless = False

            if headerless:
                df = raw_df.copy()
                df.columns = ['COD_CORTO', 'COLOR', 'TALLE', 'COD_ALFA', 'DESCRIPCION', 'CANTIDAD'] + [f'_extra_{i}' for i in range(len(df.columns) - 6)]
                df = df.dropna(how='all', axis=0)
                cod_col, qty_col, desc_col, color_col = 'COD_ALFA', 'CANTIDAD', 'DESCRIPCION', 'COLOR'
            else:
                if header_row is None:
                    header_row = 0
                df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=header_row)
                df = df.dropna(how='all', axis=1).dropna(how='all', axis=0)
                df.columns = [str(c).strip() for c in df.columns]
                cod_col = None
                for c in df.columns:
                    cu = str(c).upper().replace(' ', '').replace('_', '')
                    if 'CODALFA' in cu:
                        cod_col = c; break
                if not cod_col:
                    for c in df.columns:
                        cu = str(c).upper().replace(' ', '').replace('_', '')
                        if cu in ('ARTICULO', 'ARTÍCULO', 'ARTCULO'):
                            cod_col = c; break
                if not cod_col:
                    for c in df.columns:
                        if 'CODIGO' in str(c).upper():
                            cod_col = c; break
                qty_col = None
                for c in df.columns:
                    if 'TOTAL PARES' in str(c).upper():
                        qty_col = c; break
                if not qty_col:
                    for c in df.columns:
                        cu = str(c).upper().strip()
                        if 'CANT' in cu and 'TOTAL' in cu:
                            qty_col = c; break
                if not qty_col:
                    for c in df.columns:
                        if 'CANTIDAD' in str(c).upper() or 'PEDIDO' in str(c).upper():
                            qty_col = c; break
                if not cod_col or not qty_col:
                    hojas.append({'nombre': sheet, 'items': [], 'total_unidades': 0, 'total_items': 0, 'sin_datos': True})
                    continue
                desc_col = next((c for c in df.columns if any(k in str(c).upper() for k in ['MODELO', 'DETALLE', 'DESCRIPCION', 'DESCRIPCIÓN'])), None)
                color_col = next((c for c in df.columns if 'COLOR' in str(c).upper()), None)

            items = []
            for _, row in df.iterrows():
                if not pd.notna(row.get(cod_col)):
                    continue
                codigo = str(row[cod_col]).strip()
                if not codigo or codigo.lower() == 'nan':
                    continue
                if codigo.lower() in ('artículo', 'articulo', 'número de artículo'):
                    continue
                cantidad = 0.0
                if pd.notna(row.get(qty_col)):
                    try:
                        cantidad = float(row[qty_col])
                    except (ValueError, TypeError):
                        pass
                if cantidad <= 0:
                    continue
                cantidad_original = cantidad
                es_curva = False
                if es_miding_repo and re.search(r'[A-Za-z]\d$', codigo):
                    cantidad = cantidad * 10
                    es_curva = True
                desc = str(row.get(desc_col, '')) if desc_col else ''
                color = str(row.get(color_col, '')) if color_col else ''
                if desc.lower() == 'nan': desc = ''
                if color.lower() == 'nan': color = ''
                codigo_norm = re.sub(r'^\d{2}', '', codigo) if re.match(r'^\d{2}[A-Za-z0-9]', codigo) else codigo
                items.append({
                    'codigo': codigo,
                    'cod_normalizado': codigo_norm,
                    'modelo': desc,
                    'color': color,
                    'cantidad': cantidad,
                    'cantidad_original': cantidad_original,
                    'es_curva': es_curva,
                })
            total_u = sum(i['cantidad'] for i in items)
            total_unidades_global += total_u
            total_items_global += len(items)
            hojas.append({
                'nombre': sheet,
                'items': items,
                'total_unidades': total_u,
                'total_items': len(items),
                'sin_datos': len(items) == 0,
            })
        except Exception:
            hojas.append({'nombre': sheet, 'items': [], 'total_unidades': 0, 'total_items': 0, 'sin_datos': True})

    return {
        'solapas': hojas,
        'total_unidades': total_unidades_global,
        'total_items': total_items_global,
    }


_OMBAK_SHEETS = {'ANTIPARRA', 'CASCOS', 'GUANTES', 'TERMICOS & MEDIAS', 'ACCESORIOS', 'CAPS & HATS', 'JUNIOR & KIDS', 'CALZADO', 'INDUMENTARIA'}

_PROVIDER_KEYWORDS = {
    'OMBAK':       ['OMBAK'],
    'MIDING':      ['MIDING'],
    'MONTAGNE':    ['MONTAGNE'],
    'WORLD SPORT': ['WORLD SPORT', 'WORLDSPORT', 'WORLD-SPORT'],
}


def _detect_provider_from_filename(filename: str) -> Optional[str]:
    if not filename:
        return None
    fu = filename.upper()
    for canonical, kws in _PROVIDER_KEYWORDS.items():
        if any(k in fu for k in kws):
            return canonical
    return None


def _detect_provider_from_excel(file_bytes: bytes) -> Optional[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
        sheet_names_upper = {s.upper().strip() for s in wb.sheetnames}
        wb.close()
    except Exception:
        return None
    if len(_OMBAK_SHEETS & sheet_names_upper) >= 3:
        return 'OMBAK'
    return None


def _resolve_brand_key(*candidates: str) -> Optional[str]:
    """Devuelve la marca canónica (OMBAK|MIDING|MONTAGNE|WORLD SPORT) si alguno
    de los textos recibidos coincide con sus keywords. None si ninguna matchea."""
    for text in candidates:
        if not text:
            continue
        tu = str(text).upper()
        for canonical, kws in _PROVIDER_KEYWORDS.items():
            if any(k in tu for k in kws):
                return canonical
    return None


@router.post("/parse")
async def parse_excel_pedido(
    file: UploadFile = File(...),
    proveedor: str = Form(""),
    marca: str = Form(""),
    es_reposicion: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Parsea un Excel de nota de pedido según el contexto YA DECLARADO por el usuario
    (proveedor + marca + tipo pre/repo). NO adivina: la marca declarada manda.
    Solo como fallback (si marca/proveedor no matchean ninguna regla conocida)
    se intenta detectar por nombre de archivo o contenido.
    """
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Archivo vacío")

    # 1) Prioridad: marca declarada > proveedor declarado > filename > contenido
    declared = _resolve_brand_key(marca, proveedor)
    detected = declared or _detect_provider_from_filename(file.filename or "") or _detect_provider_from_excel(file_bytes)
    origen = "declarado" if declared else ("detectado" if detected else "desconocido")

    # Marca efectiva para elegir el parser
    brand_key = declared or detected  # None si nada matchea

    if brand_key == 'OMBAK':
        try:
            data = _parsear_excel_ombak(file_bytes)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error parseando Excel OMBAK: {e}")
    else:
        # Para el parser genérico pasamos la marca resuelta (no el nombre crudo del proveedor)
        data = _parsear_excel_generico(file_bytes, brand_key or proveedor, es_reposicion)

    # Sugerencia de provider en la BD (solo informativo)
    suggested_provider = None
    if detected and not proveedor:
        q = db.query(Provider)
        if current_user.company_id:
            q = q.filter(Provider.company_id == current_user.company_id)
        for p in q.all():
            name_u = (p.name or '').upper()
            if detected.replace(' ', '') in name_u.replace(' ', '') or detected in name_u:
                suggested_provider = {
                    'id': p.id, 'name': p.name,
                    'order_prefix': p.order_prefix, 'brands': p.brands,
                }
                break

    # Reglas aplicadas (feedback al usuario)
    rules_applied = []
    ctx = []
    if proveedor: ctx.append(f"Proveedor: {proveedor}")
    if marca:     ctx.append(f"Marca: {marca}")
    ctx.append(f"Tipo: {'REPOSICIÓN' if es_reposicion else 'PRECOMPRA'}")
    rules_applied.append("Contexto declarado → " + " · ".join(ctx))

    if brand_key == 'OMBAK':
        rules_applied.append(f"Parser OMBAK ({origen}) — multi-solapa con packs y filas ocultas")
    elif brand_key == 'MIDING':
        if es_reposicion:
            rules_applied.append(f"Parser MIDING REPOSICIÓN ({origen}) — códigos terminados en letra+dígito se multiplican ×10 (curvas)")
        else:
            rules_applied.append(f"Parser MIDING precompra ({origen}) — cantidades tal cual figuran")
    elif brand_key == 'MONTAGNE':
        rules_applied.append(f"Parser MONTAGNE ({origen}) — auto-detecta headerless en reposición")
    elif brand_key == 'WORLD SPORT':
        rules_applied.append(f"Parser WORLD SPORT ({origen}) — genérico")
    else:
        rules_applied.append("⚠ Ninguna regla específica coincide con la marca declarada — se aplicó parser genérico")

    total_curvas = sum(1 for s in data.get('solapas', []) for it in s.get('items', []) if it.get('es_curva'))
    if total_curvas > 0:
        rules_applied.append(f"{total_curvas} ítems detectados como curva (×10 aplicado)")

    return {
        'filename': file.filename,
        'proveedor_declarado': proveedor or None,
        'marca_declarada': marca or None,
        'proveedor_detectado': detected,
        'brand_key': brand_key,
        'origen_regla': origen,
        'proveedor_sugerido': suggested_provider,
        'es_reposicion': es_reposicion,
        'rules_applied': rules_applied,
        'solapas': data['solapas'],
        'total_unidades': data['total_unidades'],
        'total_items': data['total_items'],
    }
