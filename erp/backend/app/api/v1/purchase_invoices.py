"""
Router CRUD de Facturas/Remitos de Proveedor con semáforo (Purchase Invoices)
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from pydantic import BaseModel
import datetime

from app.db.session import get_db
from app.models.purchase_invoice import (
    PurchaseInvoice, PurchaseInvoiceItem,
    PurchaseInvoiceType, PurchaseInvoiceStatus, IngresoStatus, SemaforoEstado,
)
from app.models.purchase_order import PurchaseOrder
from app.models.provider import Provider
from app.models.user import User, UserRole
from app.api.deps import get_current_user, require_roles
from app.api.v1.export_utils import export_csv, export_xlsx


# ── Schemas ────────────────────────────────────────────

class PurchaseInvoiceItemOut(BaseModel):
    id: int
    code: str | None = None
    description: str | None = None
    size: str | None = None
    color: str | None = None
    quantity_invoiced: int
    quantity_received: int
    unit_price: float | None = None
    list_price: float | None = None
    model_config = {"from_attributes": True}


class PurchaseInvoiceItemCreate(BaseModel):
    code: str | None = None
    description: str | None = None
    size: str | None = None
    color: str | None = None
    quantity_invoiced: int = 0
    quantity_received: int = 0
    unit_price: float | None = None
    list_price: float | None = None


class PurchaseInvoiceOut(BaseModel):
    id: int
    number: str | None = None
    type: str
    status: str
    date: datetime.date | None = None
    due_date: datetime.date | None = None
    amount: float | None = None
    remito_venta_number: str | None = None
    linked_to_id: int | None = None
    pdf_file: str | None = None
    pdf_parsed: bool
    observations: str | None = None
    local_obs: str | None = None
    compras_obs: str | None = None
    is_partial: bool
    ingreso_status: str
    ingreso_date: datetime.datetime | None = None
    purchase_order_id: int
    purchase_order_number: str | None = None
    provider_id: int
    provider_name: str | None = None
    local_id: int | None = None
    local_name: str | None = None
    company_id: int
    created_by_id: int
    created_by_name: str | None = None
    estado_semaforo: str = "ROJO"
    confirmado_local_at: datetime.datetime | None = None
    confirmado_admin_at: datetime.datetime | None = None
    items: list[PurchaseInvoiceItemOut] = []
    model_config = {"from_attributes": True}


class PurchaseInvoiceCreate(BaseModel):
    number: str | None = None
    type: str  # FACTURA | REMITO | REMITO_FACTURA
    date: datetime.date | None = None
    due_date: datetime.date | None = None
    amount: float | None = None
    remito_venta_number: str | None = None
    observations: str | None = None
    local_obs: str | None = None
    compras_obs: str | None = None
    is_partial: bool = False
    purchase_order_id: int
    local_id: int | None = None
    items: list[PurchaseInvoiceItemCreate] = []


class PurchaseInvoiceUpdate(BaseModel):
    number: str | None = None
    type: str | None = None
    date: datetime.date | None = None
    due_date: datetime.date | None = None
    amount: float | None = None
    remito_venta_number: str | None = None
    observations: str | None = None
    local_obs: str | None = None
    compras_obs: str | None = None
    status: str | None = None
    is_partial: bool | None = None
    local_id: int | None = None
    linked_to_id: int | None = None


class IngresoConfirmBody(BaseModel):
    ingreso_date: datetime.datetime | None = None
    ingreso_photo: str | None = None
    items: list[PurchaseInvoiceItemCreate] = []  # update received quantities


class SetStatusBody(BaseModel):
    status: str  # VERDE | ROJO | ALERTA_REPO | PENDIENTE


router = APIRouter(prefix="/purchase-invoices", tags=["Purchase Invoices"])


def compute_semaforo(inv) -> str:
    """
    VERDE:    tiene RV + ingreso_status == COMPLETO + confirmado_admin_at
    AMARILLO: tiene RV + (ingreso_status PARCIAL/COMPLETO pero sin confirmado_admin)
    ROJO:     sin RV o ingreso_status == PENDIENTE/NO
    """
    if not inv.remito_venta_number:
        return SemaforoEstado.ROJO.value
    status_val = inv.ingreso_status.value if hasattr(inv.ingreso_status, "value") else inv.ingreso_status
    if status_val == "COMPLETO" and inv.confirmado_admin_at:
        return SemaforoEstado.VERDE.value
    if status_val in ("PARCIAL", "COMPLETO"):
        return SemaforoEstado.AMARILLO.value
    return SemaforoEstado.ROJO.value


# ── Serialize helper ───────────────────────────────────

def _serialize_invoice(inv: PurchaseInvoice) -> dict:
    items = [
        {
            "id": it.id,
            "code": it.code,
            "description": it.description,
            "size": it.size,
            "color": it.color,
            "quantity_invoiced": it.quantity_invoiced,
            "quantity_received": it.quantity_received,
            "unit_price": float(it.unit_price) if it.unit_price else None,
            "list_price": float(it.list_price) if it.list_price else None,
        }
        for it in inv.items
    ]
    return {
        "id": inv.id,
        "number": inv.number,
        "type": inv.type.value if inv.type else None,
        "status": inv.status.value if inv.status else None,
        "date": inv.date,
        "due_date": inv.due_date,
        "amount": float(inv.amount) if inv.amount else None,
        "remito_venta_number": inv.remito_venta_number,
        "linked_to_id": inv.linked_to_id,
        "pdf_file": inv.pdf_file,
        "pdf_parsed": inv.pdf_parsed,
        "observations": inv.observations,
        "local_obs": inv.local_obs,
        "compras_obs": inv.compras_obs,
        "is_partial": inv.is_partial,
        "ingreso_status": inv.ingreso_status.value if inv.ingreso_status else None,
        "ingreso_date": inv.ingreso_date,
        "purchase_order_id": inv.purchase_order_id,
        "purchase_order_number": inv.purchase_order.number if inv.purchase_order else None,
        "provider_id": inv.provider_id,
        "provider_name": inv.provider.name if inv.provider else None,
        "local_id": inv.local_id,
        "local_name": inv.local.name if inv.local else None,
        "company_id": inv.company_id,
        "created_by_id": inv.created_by_id,
        "created_by_name": inv.created_by.full_name if inv.created_by else None,
        "estado_semaforo": compute_semaforo(inv),
        "confirmado_local_at": inv.confirmado_local_at,
        "confirmado_admin_at": inv.confirmado_admin_at,
        "items": items,
    }


# ── Endpoints ──────────────────────────────────────────

@router.get("/")
def list_purchase_invoices(
    purchase_order_id: Optional[int] = None,
    status: Optional[str] = None,
    ingreso_status: Optional[str] = None,
    provider_id: Optional[int] = None,
    local_id: Optional[int] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=500),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    q = db.query(PurchaseInvoice)
    if current_user.role != UserRole.SUPERADMIN or current_user.company_id:
        q = q.filter(PurchaseInvoice.company_id == current_user.company_id)
    if purchase_order_id:
        q = q.filter(PurchaseInvoice.purchase_order_id == purchase_order_id)
    if status:
        try:
            q = q.filter(PurchaseInvoice.status == PurchaseInvoiceStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Status inválido: {status}")
    if ingreso_status:
        try:
            q = q.filter(PurchaseInvoice.ingreso_status == IngresoStatus(ingreso_status))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"ingreso_status inválido: {ingreso_status}")
    if provider_id:
        q = q.filter(PurchaseInvoice.provider_id == provider_id)
    if local_id:
        q = q.filter(PurchaseInvoice.local_id == local_id)
    q = q.order_by(PurchaseInvoice.date.desc(), PurchaseInvoice.id.desc())
    total = q.count()
    invoices = q.offset(skip).limit(limit).all()
    return {
        "items": [_serialize_invoice(inv) for inv in invoices],
        "total": total,
        "skip": skip,
        "limit": limit,
    }


@router.get("/export")
def export_purchase_invoices(
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    provider_id: Optional[int] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export purchase invoices as CSV or XLSX."""
    q = db.query(PurchaseInvoice)
    if current_user.role != UserRole.SUPERADMIN or current_user.company_id:
        q = q.filter(PurchaseInvoice.company_id == current_user.company_id)
    if provider_id:
        q = q.filter(PurchaseInvoice.provider_id == provider_id)
    if status:
        try:
            q = q.filter(PurchaseInvoice.status == PurchaseInvoiceStatus(status))
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Status inválido: {status}")
    q = q.order_by(PurchaseInvoice.date.desc(), PurchaseInvoice.id.desc())
    invoices = q.all()

    data = []
    for inv in invoices:
        data.append({
            "numero": inv.number or "",
            "tipo": inv.type.value if inv.type else "",
            "proveedor": inv.provider.name if inv.provider else "",
            "fecha": str(inv.date) if inv.date else "",
            "monto": float(inv.amount) if inv.amount else 0,
            "nota_pedido": inv.purchase_order.number if inv.purchase_order else "",
            "rv": inv.remito_venta_number or "",
            "estado_semaforo": inv.status.value if inv.status else "",
            "estado_ingreso": inv.ingreso_status.value if inv.ingreso_status else "",
        })

    columns = ["numero", "tipo", "proveedor", "fecha", "monto", "nota_pedido", "rv", "estado_semaforo", "estado_ingreso"]
    headers = ["Número", "Tipo", "Proveedor", "Fecha", "Monto", "Nota de Pedido", "RV", "Estado Semáforo", "Estado Ingreso"]

    if format == "csv":
        return export_csv(data, "facturas_proveedor", columns)
    return export_xlsx(data, "facturas_proveedor", columns, headers)


@router.get("/stats")
def purchase_invoices_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Estadísticas de facturas de compra."""
    company_id = current_user.company_id if current_user.role != UserRole.SUPERADMIN else None
    ci = [PurchaseInvoice.company_id == company_id] if company_id else []

    total = db.query(func.count(PurchaseInvoice.id)).filter(*ci).scalar() or 0

    # Count by semaphore status
    status_rows = (
        db.query(PurchaseInvoice.status, func.count(PurchaseInvoice.id))
        .filter(*ci)
        .group_by(PurchaseInvoice.status)
        .all()
    )
    por_status = {r[0].value: r[1] for r in status_rows}

    # Count by ingreso_status
    ingreso_rows = (
        db.query(PurchaseInvoice.ingreso_status, func.count(PurchaseInvoice.id))
        .filter(*ci)
        .group_by(PurchaseInvoice.ingreso_status)
        .all()
    )
    por_ingreso = {r[0].value: r[1] for r in ingreso_rows}

    # sin_rv
    sin_rv = (
        db.query(func.count(PurchaseInvoice.id))
        .filter(PurchaseInvoice.remito_venta_number == None, *ci)
        .scalar() or 0
    )

    # tiempos_promedio_por_proveedor: avg days between PO.date and Invoice.date
    tiempo_rows = (
        db.query(
            Provider.id,
            Provider.name,
            func.avg(
                func.extract('epoch', func.cast(PurchaseInvoice.date, func.DATE))
                - func.extract('epoch', func.cast(PurchaseOrder.date, func.DATE))
            ).label('avg_seconds'),
            func.count(PurchaseInvoice.id).label('total_facturas'),
        )
        .join(PurchaseOrder, PurchaseInvoice.purchase_order_id == PurchaseOrder.id)
        .join(Provider, PurchaseInvoice.provider_id == Provider.id)
        .filter(
            PurchaseInvoice.date != None,
            PurchaseOrder.date != None,
            *ci,
        )
        .group_by(Provider.id, Provider.name)
        .order_by(func.count(PurchaseInvoice.id).desc())
        .all()
    )
    tiempos_promedio_por_proveedor = [
        {
            "proveedor": r.name,
            "provider_id": r.id,
            "promedio_dias": round(r.avg_seconds / 86400, 1) if r.avg_seconds else 0,
            "total_facturas": r.total_facturas,
        }
        for r in tiempo_rows
    ]

    return {
        "total": total,
        "por_status": por_status,
        "por_ingreso": por_ingreso,
        "sin_rv": sin_rv,
        "tiempos_promedio_por_proveedor": tiempos_promedio_por_proveedor,
    }


@router.post("/{invoice_id}/confirm-local", summary="Confirmar recepción (LOCAL)")
def confirm_local(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.LOCAL, UserRole.DEPOSITO, UserRole.COMPRAS
    )),
):
    """LOCAL user confirms they received this document."""
    inv = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.id == invoice_id,
        PurchaseInvoice.company_id == current_user.company_id,
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura/remito no encontrado")
    if inv.confirmado_local_at:
        raise HTTPException(status_code=400, detail="Ya fue confirmado por local")
    inv.confirmado_local_at = datetime.datetime.utcnow()
    inv.confirmado_local_by_id = current_user.id
    inv.estado_semaforo = SemaforoEstado(compute_semaforo(inv))
    db.commit()
    return {"ok": True, "message": "Confirmado por local", "confirmado_local_at": inv.confirmado_local_at}


@router.post("/{invoice_id}/confirm-admin", summary="Confirmar verificación (ADMIN)")
def confirm_admin(
    invoice_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.COMPRAS
    )),
):
    """ADMIN/COMPRAS confirms they verified this document."""
    inv = db.query(PurchaseInvoice).filter(
        PurchaseInvoice.id == invoice_id,
        PurchaseInvoice.company_id == current_user.company_id,
    ).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura/remito no encontrado")
    if not inv.confirmado_local_at:
        raise HTTPException(status_code=400, detail="Debe ser confirmado por local primero")
    if inv.confirmado_admin_at:
        raise HTTPException(status_code=400, detail="Ya fue verificado por admin")
    inv.confirmado_admin_at = datetime.datetime.utcnow()
    inv.confirmado_admin_by_id = current_user.id
    inv.ingreso_status = IngresoStatus.COMPLETO
    inv.estado_semaforo = SemaforoEstado(compute_semaforo(inv))
    db.commit()
    return {"ok": True, "message": "Verificado por admin", "confirmado_admin_at": inv.confirmado_admin_at}


@router.get("/{invoice_id}")
def get_purchase_invoice(
    invoice_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")
    return _serialize_invoice(inv)


@router.post("/", status_code=201)
def create_purchase_invoice(
    body: PurchaseInvoiceCreate,
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.COMPRAS
    )),
    db: Session = Depends(get_db),
):
    company_id = current_user.company_id
    if not company_id:
        raise HTTPException(status_code=400, detail="Usuario sin company asignada")

    try:
        inv_type = PurchaseInvoiceType(body.type)
    except ValueError:
        raise HTTPException(status_code=400, detail="Tipo debe ser FACTURA, REMITO o REMITO_FACTURA")

    # Verify purchase order belongs to same company
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == body.purchase_order_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Orden de compra no encontrada")
    if po.company_id != company_id:
        raise HTTPException(status_code=403, detail="La orden de compra no pertenece a tu empresa")

    inv = PurchaseInvoice(
        number=body.number,
        type=inv_type,
        date=body.date,
        due_date=body.due_date,
        amount=body.amount,
        remito_venta_number=body.remito_venta_number,
        observations=body.observations,
        local_obs=body.local_obs,
        is_partial=body.is_partial,
        purchase_order_id=body.purchase_order_id,
        provider_id=po.provider_id,
        local_id=body.local_id,
        company_id=company_id,
        created_by_id=current_user.id,
        status=PurchaseInvoiceStatus.PENDIENTE,
        ingreso_status=IngresoStatus.PENDIENTE,
    )
    db.add(inv)
    db.flush()

    for item in body.items:
        db.add(PurchaseInvoiceItem(
            purchase_invoice_id=inv.id,
            code=item.code,
            description=item.description,
            size=item.size,
            color=item.color,
            quantity_invoiced=item.quantity_invoiced,
            quantity_received=item.quantity_received,
            unit_price=item.unit_price,
            list_price=item.list_price,
        ))

    db.commit()
    db.refresh(inv)
    return _serialize_invoice(inv)


@router.put("/{invoice_id}")
def update_purchase_invoice(
    invoice_id: int,
    body: PurchaseInvoiceUpdate,
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.COMPRAS, UserRole.LOCAL
    )),
    db: Session = Depends(get_db),
):
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")

    update_data = body.model_dump(exclude_unset=True)

    # LOCAL role can only update local_obs
    if current_user.role == UserRole.LOCAL:
        allowed = {"local_obs"}
        forbidden = set(update_data.keys()) - allowed
        if forbidden:
            raise HTTPException(
                status_code=403,
                detail=f"El rol LOCAL solo puede actualizar local_obs. Campos no permitidos: {', '.join(forbidden)}",
            )

    if "type" in update_data:
        try:
            update_data["type"] = PurchaseInvoiceType(update_data["type"])
        except ValueError:
            raise HTTPException(status_code=400, detail="Tipo debe ser FACTURA, REMITO o REMITO_FACTURA")

    if "status" in update_data:
        try:
            update_data["status"] = PurchaseInvoiceStatus(update_data["status"])
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Status inválido: {update_data['status']}")

    for key, val in update_data.items():
        setattr(inv, key, val)

    db.commit()
    db.refresh(inv)
    return _serialize_invoice(inv)


@router.post("/{invoice_id}/confirm-ingreso")
def confirm_ingreso(
    invoice_id: int,
    body: IngresoConfirmBody,
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.COMPRAS, UserRole.DEPOSITO
    )),
    db: Session = Depends(get_db),
):
    """Confirma la recepción física en depósito. Actualiza ingreso_status y quantities_received."""
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")
    if inv.ingreso_status == IngresoStatus.COMPLETO:
        raise HTTPException(status_code=400, detail="El ingreso ya fue confirmado")

    inv.ingreso_status = IngresoStatus.COMPLETO
    inv.ingreso_date = body.ingreso_date or datetime.datetime.utcnow()

    # Update quantity_received on items if provided
    if body.items:
        item_map = {it.id: it for it in inv.items}
        for item_data in body.items:
            # Match by position if no id; match by code+size+color if available
            matched = next(
                (
                    it for it in inv.items
                    if it.code == item_data.code
                    and it.size == item_data.size
                    and it.color == item_data.color
                ),
                None,
            )
            if matched:
                matched.quantity_received = item_data.quantity_received

    db.commit()
    db.refresh(inv)
    return _serialize_invoice(inv)


@router.post("/{invoice_id}/auto-link")
def auto_link_invoice(
    invoice_id: int,
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.COMPRAS
    )),
    db: Session = Depends(get_db),
):
    """Vincula automáticamente por remito_venta_number."""
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")
    if not inv.remito_venta_number:
        raise HTTPException(status_code=400, detail="La factura no tiene remito_venta_number definido")

    match = (
        db.query(PurchaseInvoice)
        .filter(
            PurchaseInvoice.company_id == inv.company_id,
            PurchaseInvoice.remito_venta_number == inv.remito_venta_number,
            PurchaseInvoice.id != invoice_id,
        )
        .first()
    )
    if not match:
        return {"linked": False, "detail": "No se encontró otra factura con el mismo remito_venta_number"}

    inv.linked_to_id = match.id
    db.commit()
    db.refresh(inv)
    return {"linked": True, "linked_to_id": match.id, "invoice": _serialize_invoice(inv)}


@router.post("/{invoice_id}/set-status")
def set_semaphore_status(
    invoice_id: int,
    body: SetStatusBody,
    current_user: User = Depends(require_roles(
        UserRole.SUPERADMIN, UserRole.ADMIN, UserRole.COMPRAS
    )),
    db: Session = Depends(get_db),
):
    """Establece manualmente el estado del semáforo (VERDE/ROJO/ALERTA_REPO/PENDIENTE)."""
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")

    try:
        inv.status = PurchaseInvoiceStatus(body.status)
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Status inválido: {body.status}. Válidos: VERDE, ROJO, ALERTA_REPO, PENDIENTE",
        )

    db.commit()
    db.refresh(inv)
    return _serialize_invoice(inv)


@router.get("/cruce")
def cruce_documentos(
    purchase_order_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Compara todos los ítems de las facturas/remitos vinculados a una Nota de Pedido.
    Agrupa por código+talle+color y suma cantidades, mostrando diferencias.
    """
    # Fetch the purchase order
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == purchase_order_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Nota de pedido no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and po.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta nota de pedido")

    company_id = po.company_id

    # Fetch all invoices linked to this purchase order
    invoices = (
        db.query(PurchaseInvoice)
        .filter(
            PurchaseInvoice.purchase_order_id == purchase_order_id,
            PurchaseInvoice.company_id == company_id,
        )
        .order_by(PurchaseInvoice.date.asc(), PurchaseInvoice.id.asc())
        .all()
    )

    # Build invoice summary list
    invoices_out = [
        {
            "id": inv.id,
            "number": inv.number,
            "type": inv.type.value if inv.type else None,
            "date": inv.date,
            "amount": float(inv.amount) if inv.amount else None,
        }
        for inv in invoices
    ]

    # Group items by (code, size, color)
    groups: dict[tuple, dict] = {}
    for inv in invoices:
        for it in inv.items:
            key = (it.code or "", it.size or "", it.color or "")
            if key not in groups:
                groups[key] = {
                    "code": it.code,
                    "description": it.description,
                    "size": it.size,
                    "color": it.color,
                    "total_invoiced": 0,
                    "total_received": 0,
                    "unit_price": None,
                    "invoice_numbers": [],
                }
            groups[key]["total_invoiced"] += it.quantity_invoiced or 0
            groups[key]["total_received"] += it.quantity_received or 0
            if it.unit_price is not None:
                groups[key]["unit_price"] = float(it.unit_price)
            if it.description and not groups[key]["description"]:
                groups[key]["description"] = it.description
            inv_num = inv.number or f"ID:{inv.id}"
            if inv_num not in groups[key]["invoice_numbers"]:
                groups[key]["invoice_numbers"].append(inv_num)

    # Compute status per item
    items_out = []
    for grp in groups.values():
        total_inv = grp["total_invoiced"]
        total_rec = grp["total_received"]
        if total_rec == 0:
            status = "SIN_RECEPCION"
        elif total_rec >= total_inv:
            status = "OK"
        else:
            status = "CON_DIFERENCIA"
        items_out.append({**grp, "status": status})

    # Sort: CON_DIFERENCIA first, then SIN_RECEPCION, then OK
    order_map = {"CON_DIFERENCIA": 0, "SIN_RECEPCION": 1, "OK": 2}
    items_out.sort(key=lambda x: (order_map.get(x["status"], 9), x["code"] or ""))

    # Summary
    total_invoiced_qty = sum(i["total_invoiced"] for i in items_out)
    total_received_qty = sum(i["total_received"] for i in items_out)
    items_ok = sum(1 for i in items_out if i["status"] == "OK")
    items_dif = sum(1 for i in items_out if i["status"] == "CON_DIFERENCIA")
    items_sin = sum(1 for i in items_out if i["status"] == "SIN_RECEPCION")

    return {
        "purchase_order_id": purchase_order_id,
        "purchase_order_number": po.number if po else None,
        "provider_name": po.provider.name if po and po.provider else None,
        "invoices": invoices_out,
        "items": items_out,
        "summary": {
            "total_items": len(items_out),
            "items_ok": items_ok,
            "items_con_diferencia": items_dif,
            "items_sin_recepcion": items_sin,
            "total_invoiced": total_invoiced_qty,
            "total_received": total_received_qty,
        },
    }


@router.delete("/{invoice_id}", status_code=204)
def delete_purchase_invoice(
    invoice_id: int,
    current_user: User = Depends(require_roles(UserRole.SUPERADMIN, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    inv = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not inv:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")

    deletable_statuses = {PurchaseInvoiceStatus.PENDIENTE, PurchaseInvoiceStatus.ROJO}
    if inv.status not in deletable_statuses:
        raise HTTPException(
            status_code=400,
            detail="Solo se puede eliminar una factura en estado PENDIENTE o ROJO",
        )

    db.delete(inv)
    db.commit()


# ── Exportar items a Excel ─────────────────────────────

@router.get("/{invoice_id}/export-items-excel")
def export_items_excel(
    invoice_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Descarga un Excel con los ítems de la factura/remito (PurchaseInvoiceItem)."""
    import io
    from fastapi.responses import Response
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado")

    inv = db.get(PurchaseInvoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")
    items = list(inv.items or [])
    if not items:
        raise HTTPException(status_code=404, detail="Este documento no tiene items cargados")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["Código", "Descripción", "Talle", "Color", "Cant. facturada", "Cant. recibida", "Precio unit.", "Precio lista"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_fact = 0
    total_rec = 0
    alt_fill = PatternFill("solid", fgColor="EFF6FF")
    for row_idx, it in enumerate(items, 2):
        qf = int(it.quantity_invoiced or 0)
        qr = int(it.quantity_received or 0)
        ws.cell(row=row_idx, column=1, value=it.code or "")
        ws.cell(row=row_idx, column=2, value=it.description or "")
        ws.cell(row=row_idx, column=3, value=it.size or "")
        ws.cell(row=row_idx, column=4, value=it.color or "")
        ws.cell(row=row_idx, column=5, value=qf)
        ws.cell(row=row_idx, column=6, value=qr)
        ws.cell(row=row_idx, column=7, value=float(it.unit_price) if it.unit_price is not None else None)
        ws.cell(row=row_idx, column=8, value=float(it.list_price) if it.list_price is not None else None)
        total_fact += qf
        total_rec += qr
        if row_idx % 2 == 0:
            for c in range(1, 9):
                ws.cell(row=row_idx, column=c).fill = alt_fill

    total_row = len(items) + 2
    total_font = Font(bold=True)
    ws.cell(row=total_row, column=4, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=5, value=total_fact).font = total_font
    ws.cell(row=total_row, column=6, value=total_rec).font = total_font

    widths = [22, 55, 10, 18, 15, 15, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    tipo = inv.type.value if inv.type else "DOC"
    numero = inv.number or str(invoice_id)
    ws.insert_rows(1)
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"{tipo} — {numero}"
    title_cell.fill = PatternFill("solid", fgColor="1E3A5F")
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_numero = numero.replace("/", "-").replace(" ", "_")
    filename = f"items_{safe_numero}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Forzar ingreso (admin) ─────────────────────────────

@router.patch("/{invoice_id}/forzar-ingreso")
def forzar_ingreso(
    invoice_id: int,
    observaciones: str = "",
    current_user: User = Depends(require_roles(UserRole.SUPERADMIN, UserRole.ADMIN)),
    db: Session = Depends(get_db),
):
    """Fuerza el ingreso desde admin: marca semáforo VERDE + confirmado admin + ingreso COMPLETO.
    Acepta `observaciones` como query param para auditoría. El motivo se concatena a
    `compras_obs` con prefijo estándar "FORZADO, SIN IMAGEN DE CARTA DE PORTE".
    """
    inv = db.get(PurchaseInvoice, invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Factura/Remito no encontrado")
    if (
        current_user.role != UserRole.SUPERADMIN
        and current_user.company_id
        and inv.company_id != current_user.company_id
    ):
        raise HTTPException(status_code=403, detail="Sin acceso a esta factura")

    now = datetime.datetime.utcnow()
    inv.estado_semaforo = SemaforoEstado.VERDE
    inv.ingreso_status = IngresoStatus.COMPLETO
    inv.ingreso_date = inv.ingreso_date or now
    inv.confirmado_admin_at = now
    inv.confirmado_admin_by_id = current_user.id

    obs_base = "FORZADO, SIN IMAGEN DE CARTA DE PORTE"
    obs_extra = (observaciones or "").strip()
    obs_final = f"{obs_base}. {obs_extra}" if obs_extra else obs_base
    prev = (inv.compras_obs or "").strip()
    inv.compras_obs = f"{prev}\n[{now.strftime('%Y-%m-%d %H:%M')}] {obs_final}".strip()

    db.commit()
    db.refresh(inv)
    return {
        "ok": True,
        "factura_id": invoice_id,
        "estado_semaforo": inv.estado_semaforo.value,
        "ingreso_status": inv.ingreso_status.value,
        "observaciones": obs_final,
    }

