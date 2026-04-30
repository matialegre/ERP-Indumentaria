"""
Módulo Reorganizador de PDF de Inventario
Procesa PDFs tipo "Inbound" de MercadoLibre y los reorganiza por Modelo → Color → Talle.
Genera respuesta JSON + descarga de Excel o PDF.
"""

import io
import re
import tempfile
import os
from collections import defaultdict
from typing import Optional

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse

import pdfplumber
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from app.api.deps import get_current_user
from app.models.user import User

router = APIRouter(prefix="/pdf-inventario", tags=["PDF Inventario"])

KNOWN_COLORS = [
    'Negro', 'Negra', 'Blanco', 'Blanca', 'Gris', 'Grafito', 'Grafiito',
    'Azul', 'Azul Marino', 'Navy', 'Petroleo', 'Petróleo', 'Acero',
    'Rojo', 'Bordo', 'Coral', 'Rosa', 'Fucsia',
    'Verde', 'Verde Oscuro', 'Militar', 'Mint',
    'Amarillo', 'Naranja',
    'Violeta', 'Grape', 'Purple',
    'Marrón Claro', 'Marrón', 'Marron', 'Chocolate', 'Caramelo', 'Carbon',
    'Beige', 'Taupe', 'Arena', 'Crema',
    'Microrayado', 'Lisa', 'Liso'
]


def extract_data_from_pdf(pdf_bytes: bytes) -> list:
    products = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                header = table[0]
                if 'PRODUCTO' not in str(header):
                    continue
                for row in table[1:]:
                    if not row or len(row) < 2:
                        continue
                    producto_text = row[0] if row[0] else ''
                    unidades_text = row[1] if len(row) > 1 and row[1] else '1'
                    if not producto_text or not producto_text.strip():
                        continue
                    try:
                        quantity = int(unidades_text.strip())
                    except Exception:
                        quantity = 1
                    products.append({'raw_text': producto_text, 'quantity': quantity})
    return products


def parse_product_info(product_text: str) -> dict:
    lines = [line.strip() for line in product_text.split('\n') if line.strip()]
    sku = ''
    product_name_parts = []
    color = ''
    size = ''
    sku_line_idx = -1

    for i, line in enumerate(lines):
        if 'SKU:' in line:
            sku_match = re.search(r'SKU:\s*([A-Z0-9\-]+)', line)
            if sku_match and sku_match.group(1) != '-':
                sku = sku_match.group(1).strip().replace('-', '')
            sku_line_idx = i
            if i + 1 < len(lines):
                next_line = lines[i + 1]
                if re.match(r'^[A-Z0-9\-]+$', next_line) and len(next_line) > 8:
                    sku = next_line.replace('-', '')
                    sku_line_idx = i + 1
            break

    last_line = lines[-1] if lines else ''
    color_size_line_idx = len(lines) - 1

    if '|' in last_line:
        parts = last_line.split('|')
        color = parts[0].strip()
        size_part = parts[1].strip() if len(parts) > 1 else ''
        size_match = re.search(r'(\d{2}(?:\.\d)?)', size_part)
        if size_match:
            size = size_match.group(1)
        else:
            size_match = re.search(r'\b(2XL|3XL|XS|XL|XXL|[SMLX]{1,2})\b', size_part)
            if size_match:
                size = size_match.group(1)
    else:
        size_match = re.search(r'(\d{2}(?:\.\d)?)\s*(?:AR|Ar)\s*$', last_line)
        if size_match:
            size = size_match.group(1)
            for known_color in sorted(KNOWN_COLORS, key=len, reverse=True):
                pattern = rf'\b({re.escape(known_color)}(?:/\w+)?)\s+{size}'
                match = re.search(pattern, last_line, re.IGNORECASE)
                if match:
                    color = match.group(1)
                    break

    for idx, line in enumerate(lines):
        if idx <= sku_line_idx:
            continue
        if idx == color_size_line_idx:
            continue
        if 'Código ML:' in line or 'Código universal:' in line or 'SKU:' in line:
            continue
        if re.match(r'^[A-Z0-9\-]+$', line) and len(line) > 8:
            continue
        product_name_parts.append(line)

    product_name = ' '.join(product_name_parts).strip()

    color_clean = color
    for prefix in ['Trekking', 'Trekkin', 'Mountain', 'Moutain', 'Urbano', 'Urbana']:
        if color_clean.lower().startswith(prefix.lower()):
            color_clean = color_clean[len(prefix):].strip()
    color_clean = color_clean.strip()

    if not color_clean:
        color_match = re.search(r'\bColor\s+(\w+)', product_name, re.IGNORECASE)
        color_clean = color_match.group(1) if color_match else 'Sin especificar'

    color_clean = color_clean.title()
    if not size:
        size = 'Único'

    return {
        'sku': sku,
        'model': product_name if product_name else 'Sin nombre',
        'color': color_clean,
        'size': size,
    }


def organize_data(products: list) -> dict:
    organized = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for product in products:
        info = parse_product_info(product['raw_text'])
        organized[info['model']][info['color']][info['size']].append({
            'sku': info['sku'],
            'quantity': product['quantity'],
        })
    return organized


def size_sort_key(s: str):
    try:
        return (0, float(s))
    except ValueError:
        order = {'XS': 0, 'S': 1, 'M': 2, 'L': 3, 'XL': 4, 'XXL': 5, '2XL': 5, '3XL': 6}
        return (1, order.get(s.upper(), 99))


def to_flat_rows(organized: dict) -> list:
    rows = []
    for model in sorted(organized.keys(), key=str.casefold):
        for color in sorted(organized[model].keys(), key=str.casefold):
            for size in sorted(organized[model][color].keys(), key=size_sort_key):
                for item in organized[model][color][size]:
                    rows.append({
                        'modelo': model,
                        'color': color,
                        'talle': size,
                        'sku': item['sku'],
                        'cantidad': item['quantity'],
                    })
    return rows


def build_excel_bytes(rows: list) -> bytes:
    df = pd.DataFrame(rows, columns=['modelo', 'color', 'talle', 'sku', 'cantidad'])
    df.columns = ['MODELO', 'COLOR', 'TALLE', 'SKU', 'CANTIDAD']

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
        ws = writer.sheets['Inventario']
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 12

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    buf.seek(0)
    return buf.read()


def build_pdf_bytes(rows: list, filename: str = '') -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15, leftMargin=15,
                            topMargin=25, bottomMargin=25)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'],
        fontSize=14, textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=12, alignment=1, fontName='Helvetica-Bold'
    )
    subtitle = f" — {filename}" if filename else ''
    elements.append(Paragraph(f"INVENTARIO ORGANIZADO - MUNDO OUTDOOR{subtitle}", title_style))
    elements.append(Spacer(1, 0.1 * inch))

    table_data = [['MODELO', 'COLOR', 'TALLE', 'SKU', 'CANT.']]
    for r in rows:
        table_data.append([r['modelo'], r['color'], r['talle'], r['sku'], str(r['cantidad'])])

    tbl = Table(table_data, colWidths=[3.5 * inch, 1.3 * inch, 0.5 * inch, 1.3 * inch, 0.4 * inch])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    elements.append(tbl)
    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.post("/process")
async def process_pdf(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """
    Recibe un PDF de inventario (ej: Inbound MercadoLibre),
    extrae productos y los organiza por Modelo → Color → Talle.
    Devuelve JSON con las filas y estadísticas.
    """
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(400, "Solo se aceptan archivos PDF")

    pdf_bytes = await file.read()
    if len(pdf_bytes) > 20 * 1024 * 1024:
        raise HTTPException(400, "El PDF no puede superar 20 MB")

    try:
        products = extract_data_from_pdf(pdf_bytes)
    except Exception as e:
        raise HTTPException(422, f"Error al leer el PDF: {str(e)}")

    if not products:
        raise HTTPException(422, "No se encontraron productos en el PDF. "
                                 "Verificá que sea un PDF de inventario con columna PRODUCTO.")

    organized = organize_data(products)
    rows = to_flat_rows(organized)

    total_qty = sum(r['cantidad'] for r in rows)
    total_models = len(organized)
    total_skus = len({r['sku'] for r in rows if r['sku']})

    return {
        "filename": file.filename,
        "rows": rows,
        "stats": {
            "total_filas": len(rows),
            "total_unidades": total_qty,
            "total_modelos": total_models,
            "total_skus": total_skus,
        }
    }


@router.post("/download-excel")
async def download_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Procesa el PDF y descarga directamente el Excel generado."""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(400, "Solo se aceptan archivos PDF")

    pdf_bytes = await file.read()
    try:
        products = extract_data_from_pdf(pdf_bytes)
        organized = organize_data(products)
        rows = to_flat_rows(organized)
        xlsx_bytes = build_excel_bytes(rows)
    except Exception as e:
        raise HTTPException(422, f"Error procesando PDF: {str(e)}")

    name = file.filename.replace('.pdf', '').replace('.PDF', '')
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="inventario_{name}.xlsx"'},
    )


@router.post("/download-pdf")
async def download_pdf_reorganizado(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Procesa el PDF y descarga el PDF reorganizado."""
    if not file.filename.lower().endswith('.pdf'):
        raise HTTPException(400, "Solo se aceptan archivos PDF")

    pdf_bytes = await file.read()
    try:
        products = extract_data_from_pdf(pdf_bytes)
        organized = organize_data(products)
        rows = to_flat_rows(organized)
        pdf_out = build_pdf_bytes(rows, file.filename)
    except Exception as e:
        raise HTTPException(422, f"Error procesando PDF: {str(e)}")

    name = file.filename.replace('.pdf', '').replace('.PDF', '')
    return StreamingResponse(
        io.BytesIO(pdf_out),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="inventario_ordenado_{name}.pdf"'},
    )
