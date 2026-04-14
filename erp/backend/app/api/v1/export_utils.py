"""Utility functions for CSV and Excel export."""
import csv
import io
from typing import List, Dict, Any
from fastapi.responses import StreamingResponse


def export_csv(data: List[Dict[str, Any]], filename: str, columns: List[str] = None) -> StreamingResponse:
    """Export data as CSV with StreamingResponse."""
    if not data:
        output = io.StringIO()
        output.write("Sin datos")
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )

    if not columns:
        columns = list(data[0].keys())

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    for row in data:
        writer.writerow(row)

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
    )


def export_xlsx(data: List[Dict[str, Any]], filename: str, columns: List[str] = None, headers: List[str] = None) -> StreamingResponse:
    """Export data as Excel (.xlsx) with StreamingResponse."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    if not columns and data:
        columns = list(data[0].keys())
    if not headers:
        headers = columns or []

    # Header row with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_key in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_key, ""))

    # Auto-width columns
    for col_idx, col_key in enumerate(columns, 1):
        max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
    )
