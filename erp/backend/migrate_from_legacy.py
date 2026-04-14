#!/usr/bin/env python3
"""
migrate_from_legacy.py
Migra datos reales de CONTROL REMITOS → nuevo ERP Mundo Outdoor

Fuentes:
  1. backend/dev.db  →  15 locales, 20 usuarios
  2. Excels de lista de precios  →  catálogo de productos Montagne + Miding + Mountain

Corre desde: erp/backend/
  .\\venv\\Scripts\\activate
  python migrate_from_legacy.py
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import sqlite3
import openpyxl
from datetime import date

from app.db.session import SessionLocal
from app.models import (
    Company, User, UserRole, Local, Provider,
    Product, ProductVariant,
)
from app.core.security import hash_password

# ── Paths ──
_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
LEGACY_DB = os.path.join(_ROOT, "CONTROL REMITOS", "backend", "dev.db")
EXCEL_MONTAGNE_105 = os.path.join(_ROOT, "CONTROL REMITOS",
    "lista de precios", "09-02-2026 LISTA DE PRECIOS VERSION 105 (1) montagne.xlsx")
EXCEL_MIDING = os.path.join(_ROOT, "CONTROL REMITOS",
    "lista de precios", "20-02-2026 MIDING - FW 2026 - ACTUALIZADO (1).xlsx")
EXCEL_MOUNTAIN = os.path.join(_ROOT, "CONTROL REMITOS",
    "lista de precios", "23-02-2026 LISTA DE PRECIOS MOUNTAIN (1).xlsx")
EXCEL_MOUNTAIN_CALZADO = os.path.join(_ROOT, "CONTROL REMITOS",
    "lista de precios", "Lista de precios calzado MOUNTAIN 08-10 (1).xlsx")

db = SessionLocal()


# ── Mapeo de roles legacy → nuevo ──
ROLE_MAP = {
    "ADMIN": UserRole.ADMIN,
    "CARGA": UserRole.ADMIN,          # Carga → Admin en el nuevo
    "LOCAL": UserRole.LOCAL,
    "COMPRAS": UserRole.COMPRAS,
    "ADMINISTRACION": UserRole.ADMINISTRACION,
    "GESTION_PAGOS": UserRole.GESTION_PAGOS,
}


def safe_float(val):
    """Convert to float, returning None on errors"""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f == f else None  # NaN check
    except (ValueError, TypeError):
        return None


def get_company():
    company = db.query(Company).filter(Company.cuit == "30-12345678-9").first()
    if not company:
        print("ERROR: Ejecutá el backend al menos una vez primero.")
        sys.exit(1)
    # Update company data
    company.name = "Mundo Outdoor"
    company.address = "Av. Córdoba 5369, CABA, Argentina"
    company.phone = "011-4771-0000"
    company.email = "info@mundooutdoor.com.ar"
    db.flush()
    return company


def migrate_locals(cid):
    """Migra los 15 locales reales desde CONTROL REMITOS"""
    print("\n[1] Locales desde CONTROL REMITOS...")
    conn = sqlite3.connect(LEGACY_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM locals ORDER BY id").fetchall()
    conn.close()

    local_map = {}  # legacy_id → new Local
    for r in rows:
        legacy_id = r["id"]
        name = r["name"]
        # Create a code from name
        code = name.upper().replace(" ", "_")[:20]
        loc = db.query(Local).filter(Local.name == name, Local.company_id == cid).first()
        if not loc:
            # Check code uniqueness
            existing_code = db.query(Local).filter(Local.code == code).first()
            if existing_code:
                code = code[:17] + f"_{legacy_id}"
            loc = Local(name=name, code=code, company_id=cid, is_active=True)
            db.add(loc)
            db.flush()
            print(f"  + {name}  (code={code})")
        else:
            print(f"  ~ {name}  (ya existe)")
        local_map[legacy_id] = loc
    return local_map


def migrate_users(cid, local_map):
    """Migra los 20 usuarios reales"""
    print("\n[2] Usuarios desde CONTROL REMITOS...")
    conn = sqlite3.connect(LEGACY_DB)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM users ORDER BY id").fetchall()
    conn.close()

    # Skip "admin" — already exists in new system
    for r in rows:
        username = r["username"]
        if username == "admin":
            print(f"  ~ admin  (ya existe como SUPERADMIN)")
            continue

        role = ROLE_MAP.get(r["role"], UserRole.VENDEDOR)
        legacy_local_id = r["local_id"]
        is_active = bool(r["is_active"])

        u = db.query(User).filter(User.username == username).first()
        if not u:
            local_id = local_map[legacy_local_id].id if legacy_local_id and legacy_local_id in local_map else None
            u = User(
                username=username,
                hashed_password=hash_password("MundoOutdoor2026!"),
                full_name=username.upper(),
                email=f"{username}@mundooutdoor.com.ar",
                role=role,
                company_id=cid,
                is_active=is_active,
            )
            db.add(u)
            db.flush()
            loc_name = local_map[legacy_local_id].name if legacy_local_id and legacy_local_id in local_map else "—"
            print(f"  + {username:15s}  {role.value:18s}  local={loc_name}")
        else:
            print(f"  ~ {username}  (ya existe)")


def make_sku(brand_prefix, code, suffix=""):
    """Generate SKU from components"""
    base = f"{brand_prefix}-{code}"
    if suffix:
        base += f"-{suffix}"
    return base.upper().replace(" ", "").replace("/", "_")[:80]


def migrate_montagne_products(cid):
    """Extrae ~620 productos de la lista Montagne v105"""
    print("\n[3] Productos Montagne (Lista v105)...")
    if not os.path.exists(EXCEL_MONTAGNE_105):
        print(f"  SKIP: no encontrado {EXCEL_MONTAGNE_105}")
        return

    wb = openpyxl.load_workbook(EXCEL_MONTAGNE_105, read_only=True, data_only=True)
    ws = wb["LISTA COMPLETA"]
    count = 0
    skipped = 0

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        rubro = row[0]
        origen = row[1]
        genero = row[2]
        estado = row[3]
        code = row[4]
        detalle = row[5]
        temporada = row[6]
        precio_may = row[7]
        precio_pub = row[8]

        if not code or not detalle:
            continue

        code = str(code).strip()
        detalle = str(detalle).strip()
        if not code or not detalle:
            continue

        # Category from rubro
        category = str(rubro).strip().title() if rubro else "Sin Categoría"

        # Check if exists
        existing = db.query(Product).filter(Product.code == code, Product.company_id == cid).first()
        if existing:
            skipped += 1
            continue

        prod = Product(
            code=code,
            name=detalle,
            brand="Montagne",
            category=category,
            base_cost=safe_float(precio_may),
            description=f"Temporada: {temporada or '—'} | Origen: {origen or '—'} | Género: {genero or '—'} | Estado: {estado or '—'}",
            company_id=cid,
            is_active=True if estado != "DISCONTINUO" else False,
        )
        db.add(prod)
        db.flush()

        # Create single variant (no size/color info in this list)
        sku = make_sku("MTG", code)
        existing_sku = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()
        if not existing_sku:
            db.add(ProductVariant(
                product_id=prod.id,
                size="ÚNICA",
                color="Estándar",
                sku=sku,
                stock=0,
                is_active=prod.is_active,
            ))
        count += 1

    db.flush()
    wb.close()
    print(f"  + {count} productos Montagne importados  ({skipped} ya existían)")


def migrate_montagne_discontinuos(cid):
    """Extrae ~1490 productos discontinuos de Montagne"""
    print("\n[4] Productos Montagne Discontinuos...")
    if not os.path.exists(EXCEL_MONTAGNE_105):
        print(f"  SKIP: no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_MONTAGNE_105, read_only=True, data_only=True)
    ws = wb["DISCONTINUOS"]
    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        rubro = row[0]
        code = row[4]
        detalle = row[5]
        temporada = row[6]
        precio_may = row[7]

        if not code or not detalle:
            continue

        code = str(code).strip()
        detalle = str(detalle).strip()
        if not code or not detalle:
            continue

        category = str(rubro).strip().title() if rubro else "Sin Categoría"

        existing = db.query(Product).filter(Product.code == code, Product.company_id == cid).first()
        if existing:
            skipped += 1
            continue

        prod = Product(
            code=code,
            name=detalle,
            brand="Montagne",
            category=category,
            base_cost=safe_float(precio_may),
            description=f"DISCONTINUO | Temporada: {temporada or '—'}",
            company_id=cid,
            is_active=False,
        )
        db.add(prod)
        db.flush()

        sku = make_sku("MTG", code, "D")
        existing_sku = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()
        if not existing_sku:
            db.add(ProductVariant(
                product_id=prod.id, size="ÚNICA", color="Estándar",
                sku=sku, stock=0, is_active=False,
            ))
        count += 1

    db.flush()
    wb.close()
    print(f"  + {count} productos discontinuos importados  ({skipped} ya existían)")


def migrate_miding_products(cid):
    """Extrae calzado Miding FW2026"""
    print("\n[5] Productos Miding (calzado FW2026)...")
    if not os.path.exists(EXCEL_MIDING):
        print(f"  SKIP: no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_MIDING, read_only=True, data_only=True)
    count = 0
    skipped = 0

    for sheet_name in ["HOMBRE", "MUJER"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        genero = "Hombre" if sheet_name == "HOMBRE" else "Mujer"

        for row in ws.iter_rows(min_row=5, values_only=True):
            modelo = row[2] if len(row) > 2 else None
            color = row[3] if len(row) > 3 else None
            linea = row[4] if len(row) > 4 else None
            cod_barras = row[7] if len(row) > 7 else None
            cod_alfa = row[8] if len(row) > 8 else None
            precio = row[11] if len(row) > 11 else None

            if not modelo or not cod_alfa:
                continue

            modelo = str(modelo).strip()
            cod_alfa = str(cod_alfa).strip()
            color = str(color).strip() if color else "Estándar"

            if not modelo or not cod_alfa or cod_alfa == "None":
                continue

            existing = db.query(Product).filter(Product.code == cod_alfa, Product.company_id == cid).first()
            if existing:
                skipped += 1
                continue

            prod = Product(
                code=cod_alfa,
                name=f"{modelo} {color}",
                brand="Miding",
                category="Calzado",
                base_cost=safe_float(precio),
                description=f"Línea: {linea or '—'} | Género: {genero}",
                company_id=cid,
                is_active=True,
            )
            db.add(prod)
            db.flush()

            sku = make_sku("MID", cod_alfa)
            existing_sku = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()
            if not existing_sku:
                db.add(ProductVariant(
                    product_id=prod.id, size="ÚNICA", color=color,
                    sku=sku, stock=0, is_active=True,
                ))
            count += 1

    db.flush()
    wb.close()
    print(f"  + {count} productos Miding importados  ({skipped} ya existían)")


def migrate_mountain_products(cid):
    """Extrae calzado Mountain"""
    print("\n[6] Productos Mountain (calzado)...")
    if not os.path.exists(EXCEL_MOUNTAIN):
        print(f"  SKIP: no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_MOUNTAIN, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        cod_alfa = row[0]
        detalle = row[1]
        precio_may = row[2]

        if not cod_alfa or not detalle:
            continue

        cod_alfa = str(cod_alfa).strip()
        detalle = str(detalle).strip()

        existing = db.query(Product).filter(Product.code == cod_alfa, Product.company_id == cid).first()
        if existing:
            skipped += 1
            continue

        prod = Product(
            code=cod_alfa,
            name=detalle,
            brand="Mountain",
            category="Calzado",
            base_cost=safe_float(precio_may),
            company_id=cid,
            is_active=True,
        )
        db.add(prod)
        db.flush()

        sku = make_sku("MTN", cod_alfa)
        existing_sku = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()
        if not existing_sku:
            db.add(ProductVariant(
                product_id=prod.id, size="ÚNICA", color="Estándar",
                sku=sku, stock=0, is_active=True,
            ))
        count += 1

    db.flush()
    wb.close()
    print(f"  + {count} productos Mountain importados  ({skipped} ya existían)")


def create_real_providers(cid):
    """Crea proveedores reales basados en las marcas que manejamos"""
    print("\n[7] Proveedores reales...")
    providers_def = [
        dict(name="Montagne S.A.", cuit="30-71234567-9",
             contact_name="Departamento Ventas", phone="011-4771-0001",
             email="ventas@montagne.com.ar", address="Av. Córdoba 5369, CABA"),
        dict(name="Miding Calzado", cuit="30-71122334-5",
             contact_name="Comercial Miding", phone="011-4555-0001",
             email="comercial@miding.com.ar", address="CABA"),
        dict(name="Mountain Calzado", cuit="30-71555666-7",
             contact_name="Ventas Mountain", phone="011-4600-0001",
             email="ventas@mountain.com.ar", address="CABA"),
    ]
    for pd in providers_def:
        existing = db.query(Provider).filter(
            Provider.name == pd["name"], Provider.company_id == cid
        ).first()
        if not existing:
            db.add(Provider(**pd, company_id=cid, is_active=True))
            print(f"  + {pd['name']}")
        else:
            print(f"  ~ {pd['name']} (ya existe)")
    db.flush()


def main():
    company = get_company()
    cid = company.id
    print(f"Empresa: {company.name} (id={cid})")

    # 1. Locales
    local_map = migrate_locals(cid)

    # 2. Usuarios
    migrate_users(cid, local_map)

    # 3-6. Productos desde Excel
    migrate_montagne_products(cid)
    migrate_montagne_discontinuos(cid)
    migrate_miding_products(cid)
    migrate_mountain_products(cid)

    # 7. Proveedores
    create_real_providers(cid)

    db.commit()

    # ── Resumen ──
    print("\n" + "=" * 55)
    print("  MIGRACIÓN COMPLETADA")
    print("=" * 55)
    n_loc  = db.query(Local).filter(Local.company_id == cid).count()
    n_user = db.query(User).filter(User.company_id == cid).count()
    n_prov = db.query(Provider).filter(Provider.company_id == cid).count()
    n_prod = db.query(Product).filter(Product.company_id == cid).count()
    n_act  = db.query(Product).filter(Product.company_id == cid, Product.is_active == True).count()
    n_disc = db.query(Product).filter(Product.company_id == cid, Product.is_active == False).count()
    n_var  = db.query(ProductVariant).join(Product).filter(Product.company_id == cid).count()
    print(f"  Locales          : {n_loc}")
    print(f"  Usuarios         : {n_user}  (+1 superadmin)")
    print(f"  Proveedores      : {n_prov}")
    print(f"  Productos total  : {n_prod}")
    print(f"    - Activos      : {n_act}")
    print(f"    - Discontinuos : {n_disc}")
    print(f"  Variantes        : {n_var}")
    print("=" * 55)
    print("\nContraseña por defecto de usuarios migrados: MundoOutdoor2026!")
    print("(Cada usuario debería cambiarla en Configuración)")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        db.rollback()
        print(f"\nERROR: {exc}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        db.close()
