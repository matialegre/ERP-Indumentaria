import pdfplumber
import pandas as pd
from collections import defaultdict
import re
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import os

def extract_data_from_pdf(pdf_path):
    """Extrae datos de las tablas del PDF"""
    products = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 2:
                    continue
                
                header = table[0]
                if 'PRODUCTO' not in str(header):
                    continue
                
                for row in table[1:]:
                    if not row or len(row) < 2:
                        continue
                    
                    producto_text = row[0] if row[0] else ''
                    unidades_text = row[1] if len(row) > 1 and row[1] else '1'
                    
                    if not producto_text or producto_text.strip() == '':
                        continue
                    
                    try:
                        quantity = int(unidades_text.strip())
                    except:
                        quantity = 1
                    
                    products.append({
                        'raw_text': producto_text,
                        'quantity': quantity
                    })
    
    return products

def parse_product_info(product_text):
    """Parsea la información del producto para extraer SKU, modelo, color y talle"""
    lines = [line.strip() for line in product_text.split('\n') if line.strip()]
    
    sku = ''
    product_name_parts = []
    color = ''
    size = ''
    
    sku_line_idx = -1
    for i, line in enumerate(lines):
        if 'SKU:' in line:
            sku_match = re.search(r'SKU:\s*([A-Z0-9\-]+)', line)
            if sku_match and sku_match.group(1) != '-':
                sku = sku_match.group(1).strip()
            sku_line_idx = i
            
            if i + 1 < len(lines):
                next_line = lines[i + 1]
                if re.match(r'^[A-Z0-9\-]+$', next_line) and len(next_line) > 8:
                    sku = next_line
                    sku_line_idx = i + 1
            break
    
    last_line = lines[-1] if lines else ''
    
    known_colors = [
        'Negro', 'Negra', 'Blanco', 'Blanca', 'Gris', 'Grafito', 'Grafiito',
        'Azul', 'Azul Marino', 'Navy', 'Petroleo', 'Petróleo', 'Acero',
        'Rojo', 'Bordo', 'Coral', 'Rosa', 'Fucsia',
        'Verde', 'Verde Oscuro', 'Militar', 'Mint',
        'Amarillo', 'Naranja',
        'Violeta', 'Grape', 'Purple',
        'Marrón Claro', 'Marrón', 'Marron', 'Chocolate', 'Caramelo', 'Carbon',
        'Beige', 'Taupe', 'Arena', 'Crema',
        'Microrayado', 'Lisa', 'Liso'
    ]
    
    color_size_line_idx = len(lines) - 1
    
    if '|' in last_line:
        parts = last_line.split('|')
        color = parts[0].strip()
        size_part = parts[1].strip() if len(parts) > 1 else ''
        
        size_match = re.search(r'(\d{2}(?:\.\d)?)', size_part)
        if size_match:
            size = size_match.group(1)
        else:
            size_match = re.search(r'\b(2XL|3XL|XS|XL|XXL|[SMLX]{1,2})\b', size_part)
            if size_match:
                size = size_match.group(1)
    else:
        size_match = re.search(r'(\d{2}(?:\.\d)?)\s*(?:AR|Ar)\s*$', last_line)
        if size_match:
            size = size_match.group(1)
            
            for known_color in sorted(known_colors, key=len, reverse=True):
                pattern = rf'\b({re.escape(known_color)}(?:/\w+)?)\s+{size}'
                match = re.search(pattern, last_line, re.IGNORECASE)
                if match:
                    color = match.group(1)
                    break
    
    for idx, line in enumerate(lines):
        if idx <= sku_line_idx:
            continue
        if idx == color_size_line_idx:
            continue
        if 'Código ML:' in line or 'Código universal:' in line or 'SKU:' in line:
            continue
        if re.match(r'^[A-Z0-9\-]+$', line) and len(line) > 8:
            continue
        
        product_name_parts.append(line)
    
    product_name = ' '.join(product_name_parts).strip()
    
    prefixes_to_remove = ['Trekking', 'Trekkin', 'Mountain', 'Moutain', 'Urbano', 'Urbana']
    color_clean = color
    for prefix in prefixes_to_remove:
        if color_clean.lower().startswith(prefix.lower()):
            color_clean = color_clean[len(prefix):].strip()
    
    color_clean = color_clean.strip()
    
    if not color_clean:
        color_match = re.search(r'\bColor\s+(\w+)', product_name, re.IGNORECASE)
        if color_match:
            color_clean = color_match.group(1)
        else:
            color_clean = 'Sin especificar'
    
    if not size:
        size = 'Único'
    
    color_clean = color_clean.title()
    
    return {
        'sku': sku,
        'model': product_name if product_name else 'Sin nombre',
        'color': color_clean,
        'size': size
    }

def organize_data(products):
    """Organiza los productos por MODELO -> COLOR -> TALLE"""
    organized = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for product in products:
        info = parse_product_info(product['raw_text'])
        
        organized[info['model']][info['color']][info['size']].append({
            'sku': info['sku'],
            'quantity': product['quantity'],
            'raw_text': product['raw_text']
        })
    
    return organized

def create_excel(organized_data, output_path):
    """Crea un archivo Excel ordenado por MODELO -> COLOR -> TALLE"""
    rows = []
    
    for model in sorted(organized_data.keys()):
        for color in sorted(organized_data[model].keys()):
            for size in sorted(organized_data[model][color].keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else x):
                items = organized_data[model][color][size]
                for item in items:
                    rows.append({
                        'MODELO': model,
                        'COLOR': color,
                        'TALLE': size,
                        'SKU': item['sku'],
                        'CANTIDAD': item['quantity']
                    })
    
    df = pd.DataFrame(rows)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
        
        worksheet = writer.sheets['Inventario']
        
        worksheet.column_dimensions['A'].width = 60
        worksheet.column_dimensions['B'].width = 20
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 20
        worksheet.column_dimensions['E'].width = 12
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    print(f"✓ Excel creado: {output_path}")
    print(f"  Total de items: {len(rows)}")
    return df

def create_pdf(organized_data, output_path):
    """Crea un PDF formateado para imprimir, todo concatenado y organizado por MODELO -> COLOR -> TALLE"""
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                           rightMargin=20, leftMargin=20,
                           topMargin=30, bottomMargin=30)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=20,
        alignment=1,
        fontName='Helvetica-Bold'
    )
    
    model_style = ParagraphStyle(
        'ModelStyle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=6,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    color_style = ParagraphStyle(
        'ColorStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=4,
        spaceBefore=6,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("INVENTARIO ORGANIZADO - MUNDO OUTDOOR", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    all_table_data = [['MODELO', 'COLOR', 'TALLE', 'SKU', 'CANT.']]
    
    for model in sorted(organized_data.keys()):
        for color in sorted(organized_data[model].keys()):
            for size in sorted(organized_data[model][color].keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else x):
                items = organized_data[model][color][size]
                for item in items:
                    model_short = model[:50] + '...' if len(model) > 50 else model
                    all_table_data.append([
                        model_short,
                        color,
                        size,
                        item['sku'],
                        str(item['quantity'])
                    ])
    
    table = Table(all_table_data, colWidths=[3.2*inch, 1.2*inch, 0.6*inch, 1.3*inch, 0.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    print(f"✓ PDF creado: {output_path}")

def main():
    """Función principal"""
    print("=" * 70)
    print("    REORGANIZADOR DE INVENTARIO - MUNDO OUTDOOR")
    print("=" * 70)
    
    pdf_input = r"c:\Users\Mundo Outdoor\Desktop\pdfordenado\Inbound-61958574-preparation-instructions (2) (1).pdf"
    excel_output = r"c:\Users\Mundo Outdoor\Desktop\pdfordenado\inventario_ordenado.xlsx"
    pdf_output = r"c:\Users\Mundo Outdoor\Desktop\pdfordenado\inventario_ordenado.pdf"
    
    if not os.path.exists(pdf_input):
        print(f"❌ Error: No se encuentra el archivo {pdf_input}")
        return
    
    print(f"\n📄 Procesando: {os.path.basename(pdf_input)}\n")
    
    print("[1/4] Extrayendo productos de las tablas del PDF...")
    products = extract_data_from_pdf(pdf_input)
    print(f"      ✓ {len(products)} productos extraídos")
    
    print("\n[2/4] Organizando por MODELO → COLOR → TALLE...")
    organized_data = organize_data(products)
    total_models = len(organized_data)
    total_colors = sum(len(colors) for colors in organized_data.values())
    print(f"      ✓ {total_models} modelos diferentes")
    print(f"      ✓ {total_colors} combinaciones modelo-color")
    
    print("\n[3/4] Generando archivo Excel...")
    df = create_excel(organized_data, excel_output)
    
    print("\n[4/4] Generando PDF concatenado para imprimir...")
    create_pdf(organized_data, pdf_output)
    
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 70)
    print(f"\n📊 Archivos generados:")
    print(f"   • Excel: {excel_output}")
    print(f"   • PDF:   {pdf_output}")
    print(f"\n📈 Estadísticas:")
    print(f"   • Productos totales: {len(df)}")
    print(f"   • Modelos únicos: {total_models}")
    print(f"   • Combinaciones modelo-color: {total_colors}")
    
    total_qty = df['CANTIDAD'].sum()
    print(f"   • Unidades totales: {total_qty}")
    print("\n✨ Los archivos están listos para usar!")
    print("   - Excel: Para editar y trabajar con los datos")
    print("   - PDF: Para imprimir (todo concatenado en una tabla)\n")

if __name__ == "__main__":
    main()
