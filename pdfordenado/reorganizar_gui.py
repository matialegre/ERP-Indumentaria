import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
import pdfplumber
import pandas as pd
import re
from collections import defaultdict
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def extract_data_from_pdf(pdf_path):
    """Extrae datos de las tablas del PDF"""
    products = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 2:
                    continue
                
                header = table[0]
                if 'PRODUCTO' not in str(header):
                    continue
                
                for row in table[1:]:
                    if not row or len(row) < 2:
                        continue
                    
                    producto_text = row[0] if row[0] else ''
                    unidades_text = row[1] if len(row) > 1 and row[1] else '1'
                    
                    if not producto_text or producto_text.strip() == '':
                        continue
                    
                    try:
                        quantity = int(unidades_text.strip())
                    except:
                        quantity = 1
                    
                    products.append({
                        'raw_text': producto_text,
                        'quantity': quantity
                    })
    
    return products


def parse_product_info(product_text):
    """Parsea la información del producto para extraer SKU, modelo, color y talle"""
    lines = [line.strip() for line in product_text.split('\n') if line.strip()]
    
    sku = ''
    product_name_parts = []
    color = ''
    size = ''
    
    sku_line_idx = -1
    for i, line in enumerate(lines):
        if 'SKU:' in line:
            sku_match = re.search(r'SKU:\s*([A-Z0-9\-]+)', line)
            if sku_match and sku_match.group(1) != '-':
                sku = sku_match.group(1).strip()
            sku_line_idx = i
            
            if i + 1 < len(lines):
                next_line = lines[i + 1]
                if re.match(r'^[A-Z0-9\-]+$', next_line) and len(next_line) > 8:
                    sku = next_line
                    sku_line_idx = i + 1
            break
    
    last_line = lines[-1] if lines else ''
    
    known_colors = [
        'Negro', 'Negra', 'Blanco', 'Blanca', 'Gris', 'Grafito', 'Grafiito',
        'Azul', 'Azul Marino', 'Navy', 'Petroleo', 'Petróleo', 'Acero',
        'Rojo', 'Bordo', 'Coral', 'Rosa', 'Fucsia',
        'Verde', 'Verde Oscuro', 'Militar', 'Mint',
        'Amarillo', 'Naranja',
        'Violeta', 'Grape', 'Purple',
        'Marrón Claro', 'Marrón', 'Marron', 'Chocolate', 'Caramelo', 'Carbon',
        'Beige', 'Taupe', 'Arena', 'Crema',
        'Microrayado', 'Lisa', 'Liso'
    ]
    
    color_size_line_idx = len(lines) - 1
    
    if '|' in last_line:
        parts = last_line.split('|')
        color = parts[0].strip()
        size_part = parts[1].strip() if len(parts) > 1 else ''
        
        size_match = re.search(r'(\d{2}(?:\.\d)?)', size_part)
        if size_match:
            size = size_match.group(1)
        else:
            size_match = re.search(r'\b(2XL|3XL|XS|XL|XXL|[SMLX]{1,2})\b', size_part)
            if size_match:
                size = size_match.group(1)
    else:
        size_match = re.search(r'(\d{2}(?:\.\d)?)\s*(?:AR|Ar)\s*$', last_line)
        if size_match:
            size = size_match.group(1)
            
            for known_color in sorted(known_colors, key=len, reverse=True):
                pattern = rf'\b({re.escape(known_color)}(?:/\w+)?)\s+{size}'
                match = re.search(pattern, last_line, re.IGNORECASE)
                if match:
                    color = match.group(1)
                    break
    
    for idx, line in enumerate(lines):
        if idx <= sku_line_idx:
            continue
        if idx == color_size_line_idx:
            continue
        if 'Código ML:' in line or 'Código universal:' in line or 'SKU:' in line:
            continue
        if re.match(r'^[A-Z0-9\-]+$', line) and len(line) > 8:
            continue
        
        product_name_parts.append(line)
    
    product_name = ' '.join(product_name_parts).strip()
    
    prefixes_to_remove = ['Trekking', 'Trekkin', 'Mountain', 'Moutain', 'Urbano', 'Urbana']
    color_clean = color
    for prefix in prefixes_to_remove:
        if color_clean.lower().startswith(prefix.lower()):
            color_clean = color_clean[len(prefix):].strip()
    
    color_clean = color_clean.strip()
    
    if not color_clean:
        color_match = re.search(r'\bColor\s+(\w+)', product_name, re.IGNORECASE)
        if color_match:
            color_clean = color_match.group(1)
        else:
            color_clean = 'Sin especificar'
    
    if not size:
        size = 'Único'
    
    color_clean = color_clean.title()
    
    return {
        'sku': sku,
        'model': product_name if product_name else 'Sin nombre',
        'color': color_clean,
        'size': size
    }


def organize_data(products):
    """Organiza los productos por MODELO -> COLOR -> TALLE"""
    organized = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    for product in products:
        info = parse_product_info(product['raw_text'])
        
        organized[info['model']][info['color']][info['size']].append({
            'sku': info['sku'],
            'quantity': product['quantity']
        })
    
    return organized


def create_excel(organized_data, output_path):
    """Crea un archivo Excel ordenado por MODELO -> COLOR -> TALLE"""
    rows = []
    
    for model in sorted(organized_data.keys()):
        for color in sorted(organized_data[model].keys()):
            for size in sorted(organized_data[model][color].keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else x):
                items = organized_data[model][color][size]
                for item in items:
                    rows.append({
                        'MODELO': model,
                        'COLOR': color,
                        'TALLE': size,
                        'SKU': item['sku'],
                        'CANTIDAD': item['quantity']
                    })
    
    df = pd.DataFrame(rows)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventario', index=False)
        
        worksheet = writer.sheets['Inventario']
        
        worksheet.column_dimensions['A'].width = 70
        worksheet.column_dimensions['B'].width = 22
        worksheet.column_dimensions['C'].width = 10
        worksheet.column_dimensions['D'].width = 22
        worksheet.column_dimensions['E'].width = 12
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    return df


def create_pdf(organized_data, output_path):
    """Crea un PDF formateado para imprimir"""
    doc = SimpleDocTemplate(output_path, pagesize=A4,
                           rightMargin=15, leftMargin=15,
                           topMargin=25, bottomMargin=25)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=15,
        alignment=1,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("INVENTARIO ORGANIZADO - MUNDO OUTDOOR", title_style))
    elements.append(Spacer(1, 0.15*inch))
    
    all_table_data = [['MODELO', 'COLOR', 'TALLE', 'SKU', 'CANT.']]
    
    for model in sorted(organized_data.keys()):
        for color in sorted(organized_data[model].keys()):
            for size in sorted(organized_data[model][color].keys(), key=lambda x: float(x) if x.replace('.','').isdigit() else x):
                items = organized_data[model][color][size]
                for item in items:
                    all_table_data.append([
                        model,
                        color,
                        size,
                        item['sku'],
                        str(item['quantity'])
                    ])
    
    table = Table(all_table_data, colWidths=[3.5*inch, 1.3*inch, 0.5*inch, 1.3*inch, 0.4*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 0), (4, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))
    
    elements.append(table)
    doc.build(elements)


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Reorganizador de Inventario - Mundo Outdoor")
        self.root.geometry("600x450")
        self.root.resizable(False, False)
        
        # Colores
        self.bg_color = "#f0f4f8"
        self.accent_color = "#2c3e50"
        self.button_color = "#3498db"
        
        self.root.configure(bg=self.bg_color)
        
        # Frame principal
        main_frame = tk.Frame(root, bg=self.bg_color, padx=30, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título
        title_label = tk.Label(main_frame, text="📦 Reorganizador de Inventario",
                              font=("Segoe UI", 20, "bold"), bg=self.bg_color, fg=self.accent_color)
        title_label.pack(pady=(0, 5))
        
        subtitle_label = tk.Label(main_frame, text="Mundo Outdoor",
                                 font=("Segoe UI", 12), bg=self.bg_color, fg="#7f8c8d")
        subtitle_label.pack(pady=(0, 20))
        
        # Frame para selección de archivo
        file_frame = tk.LabelFrame(main_frame, text=" Archivo PDF ", font=("Segoe UI", 10, "bold"),
                                   bg=self.bg_color, fg=self.accent_color, padx=15, pady=15)
        file_frame.pack(fill=tk.X, pady=10)
        
        self.file_path = tk.StringVar()
        self.file_entry = tk.Entry(file_frame, textvariable=self.file_path, font=("Segoe UI", 10),
                                   width=45, state='readonly')
        self.file_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        browse_btn = tk.Button(file_frame, text="Seleccionar...", font=("Segoe UI", 10),
                              bg=self.button_color, fg="white", cursor="hand2",
                              command=self.browse_file, relief=tk.FLAT, padx=15)
        browse_btn.pack(side=tk.LEFT)
        
        # Frame para carpeta de salida
        output_frame = tk.LabelFrame(main_frame, text=" Carpeta de Salida ", font=("Segoe UI", 10, "bold"),
                                     bg=self.bg_color, fg=self.accent_color, padx=15, pady=15)
        output_frame.pack(fill=tk.X, pady=10)
        
        self.output_path = tk.StringVar()
        self.output_entry = tk.Entry(output_frame, textvariable=self.output_path, font=("Segoe UI", 10),
                                     width=45, state='readonly')
        self.output_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        output_btn = tk.Button(output_frame, text="Seleccionar...", font=("Segoe UI", 10),
                              bg=self.button_color, fg="white", cursor="hand2",
                              command=self.browse_output, relief=tk.FLAT, padx=15)
        output_btn.pack(side=tk.LEFT)
        
        # Botón procesar
        self.process_btn = tk.Button(main_frame, text="🚀 PROCESAR INVENTARIO", font=("Segoe UI", 14, "bold"),
                                    bg="#27ae60", fg="white", cursor="hand2",
                                    command=self.process, relief=tk.FLAT, padx=30, pady=12)
        self.process_btn.pack(pady=25)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(main_frame, length=400, mode='indeterminate')
        self.progress.pack(pady=5)
        
        # Estado
        self.status_var = tk.StringVar(value="Seleccione un archivo PDF para comenzar")
        self.status_label = tk.Label(main_frame, textvariable=self.status_var,
                                    font=("Segoe UI", 10), bg=self.bg_color, fg="#7f8c8d")
        self.status_label.pack(pady=10)
        
        # Resultados
        self.result_frame = tk.Frame(main_frame, bg=self.bg_color)
        self.result_frame.pack(fill=tk.X, pady=10)
    
    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Seleccionar PDF de inventario",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            # Auto-set output folder
            self.output_path.set(os.path.dirname(filename))
            self.status_var.set("✓ Archivo seleccionado. Listo para procesar.")
    
    def browse_output(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta de salida")
        if folder:
            self.output_path.set(folder)
    
    def process(self):
        if not self.file_path.get():
            messagebox.showerror("Error", "Por favor seleccione un archivo PDF")
            return
        
        if not self.output_path.get():
            messagebox.showerror("Error", "Por favor seleccione una carpeta de salida")
            return
        
        self.process_btn.config(state=tk.DISABLED)
        self.progress.start(10)
        self.status_var.set("⏳ Procesando...")
        
        thread = threading.Thread(target=self.run_process)
        thread.start()
    
    def run_process(self):
        try:
            pdf_input = self.file_path.get()
            output_folder = self.output_path.get()
            
            excel_output = os.path.join(output_folder, "inventario_ordenado.xlsx")
            pdf_output = os.path.join(output_folder, "inventario_ordenado.pdf")
            
            # Extraer datos
            self.root.after(0, lambda: self.status_var.set("⏳ Extrayendo datos del PDF..."))
            products = extract_data_from_pdf(pdf_input)
            
            # Organizar
            self.root.after(0, lambda: self.status_var.set("⏳ Organizando productos..."))
            organized_data = organize_data(products)
            
            # Crear Excel
            self.root.after(0, lambda: self.status_var.set("⏳ Generando Excel..."))
            df = create_excel(organized_data, excel_output)
            
            # Crear PDF
            self.root.after(0, lambda: self.status_var.set("⏳ Generando PDF..."))
            create_pdf(organized_data, pdf_output)
            
            # Estadísticas
            total_products = len(df)
            total_qty = df['CANTIDAD'].sum()
            total_models = len(organized_data)
            
            self.root.after(0, lambda: self.on_success(total_products, total_qty, total_models, excel_output, pdf_output))
            
        except Exception as e:
            self.root.after(0, lambda: self.on_error(str(e)))
    
    def on_success(self, total_products, total_qty, total_models, excel_path, pdf_path):
        self.progress.stop()
        self.process_btn.config(state=tk.NORMAL)
        
        self.status_var.set(f"✅ ¡Completado! {total_products} productos, {total_qty} unidades, {total_models} modelos")
        
        # Limpiar resultados anteriores
        for widget in self.result_frame.winfo_children():
            widget.destroy()
        
        result_text = f"📊 Excel: {os.path.basename(excel_path)}\n📄 PDF: {os.path.basename(pdf_path)}"
        result_label = tk.Label(self.result_frame, text=result_text, font=("Segoe UI", 9),
                               bg=self.bg_color, fg="#27ae60", justify=tk.LEFT)
        result_label.pack()
        
        # Botón para abrir carpeta
        open_btn = tk.Button(self.result_frame, text="📁 Abrir Carpeta", font=("Segoe UI", 9),
                            bg="#95a5a6", fg="white", cursor="hand2",
                            command=lambda: os.startfile(self.output_path.get()), relief=tk.FLAT, padx=10)
        open_btn.pack(pady=10)
        
        messagebox.showinfo("¡Éxito!", f"Archivos generados correctamente:\n\n• {os.path.basename(excel_path)}\n• {os.path.basename(pdf_path)}\n\nTotal: {total_products} productos, {total_qty} unidades")
    
    def on_error(self, error_msg):
        self.progress.stop()
        self.process_btn.config(state=tk.NORMAL)
        self.status_var.set("❌ Error al procesar")
        messagebox.showerror("Error", f"Ocurrió un error:\n\n{error_msg}")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
